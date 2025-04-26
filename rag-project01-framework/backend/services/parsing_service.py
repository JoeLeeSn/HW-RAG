import logging
from typing import Dict, List
import fitz  # PyMuPDF
import pandas as pd
from datetime import datetime
import markdown
import mammoth  # for docx
import openpyxl  # for excel
import pytesseract
from PIL import Image
import io
import re

logger = logging.getLogger(__name__)

class ParsingService:
    """
    PDF文档解析服务类
    
    该类提供多种解析策略来提取和构建文档内容，包括：
    - 全文提取
    - 逐页解析
    - 基于标题的分段
    - 文本和表格混合解析
    - 图片提取和OCR
    """

    def parse_document(self, file_content: bytes, file_type: str, method: str, metadata: dict) -> dict:
        """
        使用指定方法解析文档

        参数:
            file_content (bytes): 文件内容
            file_type (str): 文件类型 ('pdf', 'markdown', 'docx', 'excel')
            method (str): 解析方法
            metadata (dict): 文档元数据

        返回:
            dict: 解析后的文档数据
        """
        try:
            if file_type == 'pdf':
                return self._parse_pdf(file_content, method, metadata)
            elif file_type == 'markdown':
                return self._parse_markdown(file_content, method, metadata)
            elif file_type == 'docx':
                return self._parse_docx(file_content, method, metadata)
            elif file_type == 'excel':
                return self._parse_excel(file_content, method, metadata)
            else:
                raise ValueError(f"Unsupported file type: {file_type}")
                
        except Exception as e:
            logger.error(f"Error in parse_document: {str(e)}")
            raise

    def _parse_pdf(self, file_content: bytes, method: str, metadata: dict) -> dict:
        """解析PDF文档"""
        doc = fitz.open(stream=file_content, filetype="pdf")
        page_map = []
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            page_content = {
                "page": page_num + 1,
                "text": page.get_text(),
                "images": [],
                "tables": []
            }
            
            # 提取图片
            if method in ["extract_images", "all_text"]:
                for img_index, img in enumerate(page.get_images()):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    
                    # 使用OCR提取图片中的文字
                    image = Image.open(io.BytesIO(image_bytes))
                    ocr_text = pytesseract.image_to_string(image, lang='chi_sim+eng')
                    
                    page_content["images"].append({
                        "index": img_index,
                        "content": f"图片 {img_index + 1}",
                        "ocr_text": ocr_text.strip()
                    })
            
            # 提取表格
            if method in ["extract_tables", "all_text"]:
                tables = page.find_tables()
                for table_index, table in enumerate(tables):
                    table_data = table.extract()
                    table_markdown = self._table_to_markdown(table_data)
                    page_content["tables"].append({
                        "index": table_index,
                        "content": table_markdown
                    })
            
            page_map.append(page_content)
        
        return self._process_content(page_map, method, metadata)

    def _parse_markdown(self, file_content: bytes, method: str, metadata: dict) -> dict:
        """解析Markdown文档"""
        content = file_content.decode('utf-8')
        html = markdown.markdown(content)
        
        # 提取图片
        images = re.findall(r'!\[(.*?)\]\((.*?)\)', content)
        image_content = [{"content": desc, "url": url} for desc, url in images]
        
        # 提取表格
        tables = re.findall(r'\|.*?\|', content, re.MULTILINE)
        table_content = [{"content": table} for table in tables]
        
        return {
            "metadata": {
                **metadata,
                "file_type": "markdown",
                "parsing_method": method,
                "timestamp": datetime.now().isoformat()
            },
            "content": [
                {"type": "text", "content": content},
                *[{"type": "image", **img} for img in image_content],
                *[{"type": "table", **table} for table in table_content]
            ]
        }

    def _parse_docx(self, file_content: bytes, method: str, metadata: dict) -> dict:
        """解析Word文档"""
        result = mammoth.convert_to_markdown(io.BytesIO(file_content))
        markdown_content = result.value
        
        return self._parse_markdown(markdown_content.encode('utf-8'), method, metadata)

    def _parse_excel(self, file_content: bytes, method: str, metadata: dict) -> dict:
        """解析Excel文档"""
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        content = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            table_data = []
            
            for row in sheet.iter_rows():
                table_data.append([cell.value for cell in row])
            
            table_markdown = self._table_to_markdown(table_data)
            content.append({
                "type": "table",
                "title": sheet_name,
                "content": table_markdown
            })
        
        return {
            "metadata": {
                **metadata,
                "file_type": "excel",
                "parsing_method": method,
                "timestamp": datetime.now().isoformat()
            },
            "content": content
        }

    def _process_content(self, page_map: list, method: str, metadata: dict) -> dict:
        """处理解析后的内容"""
        parsed_content = []
        
        if method == "all_text":
            for page in page_map:
                parsed_content.append({
                    "type": "text",
                    "page": page["page"],
                    "content": page["text"]
                })
                parsed_content.extend([
                    {**img, "type": "image", "page": page["page"]} 
                    for img in page["images"]
                ])
                parsed_content.extend([
                    {**table, "type": "table", "page": page["page"]} 
                    for table in page["tables"]
                ])
        elif method == "extract_images":
            for page in page_map:
                parsed_content.extend([
                    {**img, "type": "image", "page": page["page"]} 
                    for img in page["images"]
                ])
        elif method == "extract_tables":
            for page in page_map:
                parsed_content.extend([
                    {**table, "type": "table", "page": page["page"]} 
                    for table in page["tables"]
                ])
        else:
            raise ValueError(f"Unsupported parsing method: {method}")
        
        return {
            "metadata": {
                **metadata,
                "file_type": "pdf",
                "parsing_method": method,
                "timestamp": datetime.now().isoformat()
            },
            "content": parsed_content
        }

    def _table_to_markdown(self, table_data: list) -> str:
        """将表格数据转换为Markdown格式"""
        if not table_data:
            return ""
            
        # 创建表头分隔行
        header = "| " + " | ".join(str(cell) for cell in table_data[0]) + " |"
        separator = "| " + " | ".join("---" for _ in table_data[0]) + " |"
        
        # 创建数据行
        rows = []
        for row in table_data[1:]:
            rows.append("| " + " | ".join(str(cell) for cell in row) + " |")
        
        return "\n".join([header, separator] + rows) 